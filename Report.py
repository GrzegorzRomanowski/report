import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
from pycel import ExcelCompiler

from config import config_env, Config
from interface import report_welcome_message_box, warning_msg_box


# region Initialization

# Adjusting style
# colors
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
grey_fill = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type='solid')
title_fill = PatternFill(start_color='F0E68C', end_color='F0E68C', fill_type='solid')
summary_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
pink_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
green_fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
# borders
edge = Side(border_style='thin', color='000000')
thick_edge = Side(border_style='thick', color='000000')
border = Border(top=edge, bottom=edge, left=edge, right=edge)
thick_border = Border(top=thick_edge, bottom=thick_edge, left=thick_edge, right=thick_edge)
# fonts
font_bold = Font(bold=True)

# endregion


# region Transform data
class Report:
    def __init__(self, config: Config, user_data: dict):
        self.config = config
        self.user_data = user_data
        # Load Workbooks (wb)
        try:
            self.wb_ebawe = openpyxl.load_workbook(config.EBAWE_REPORT_PATH)
        except (FileNotFoundError, NameError):
            warn = "Nie znaleziono pliku z raportem EBAWE.\n\nSkrypt zamknie się bez zapisywania żadnych zmian.\n\n" \
                   "Sprawdź czy raport jest stworzony.\nZweryfikuj jego nazwę, rozszerzenie, lokalizację itp.\n" \
                   "i uruchom skrypt ponownie."
            warning_msg_box(warn)
        if os.path.isfile(config.DAILY_TEMP_PATH):
            self.wb_daily = openpyxl.load_workbook(config.DAILY_TEMP_PATH)
        else:
            self.wb_daily = openpyxl.load_workbook(config.TEMPLATE_PATH)

        if os.path.isfile(config.YEARLY_TEMP_PATH):
            self.wb_yearly = openpyxl.load_workbook(config.YEARLY_TEMP_PATH)
        else:
            self.wb_yearly = openpyxl.load_workbook(config.YEARLY_PATH)

        if os.path.isfile(config.MONTHLY_TEMP_PATH):
            self.wb_month = openpyxl.load_workbook(config.MONTHLY_TEMP_PATH)
        else:
            self.wb_month = openpyxl.load_workbook(config.MONTHLY_PATH)

        # Load Worksheets (ws)
        self.ws_ebawe = self.wb_ebawe.active
        self.ws_daily = self.wb_daily[f"E{self.user_data['ebawe']}"]
        self.ws_yearly = self.wb_yearly.active
        self.ws_monthly = self.wb_month[f"E{self.user_data['ebawe']}"]
        # Row 1 variables
        self.row_1_yearly = self.ws_yearly[1]
        self.row_1_monthly = self.ws_monthly[1]
        # Number of rows in sheets
        self.ebawe_max_row = self.ws_ebawe.max_row
        self.yearly_max_row = self.ws_yearly.max_row
        self.yearly_max_col = self.ws_yearly.max_column

        # Columns names
        self.title_tuple = ("Element", "Typ", "Powierzchnia", "Klasa betonu", "Gatunek stali")
        self.title_index = (3, 4, 6, 7, 9)

        # List of projects
        self.ebawe_project_list = list()

    def __call__(self):
        self.fill_ebawe_project_list()
        self.create_daily_report()
        self.painting_monthly_and_yearly_reports()
        self.filling_monthly_report()
        self.saving_all_reports_files()

    def fill_ebawe_project_list(self):
        """ Fill in data in self.ebawe_project_list according to the schema below
        self.ebawe_project_list = [[proj_number, row_index_from_ebawe_raport, number_of elements], [...], ...]
        :return:
        """
        proj_counter = 0
        for row_index in range(1, self.ebawe_max_row):
            if self.ws_ebawe.cell(column=5, row=row_index).value is not None:
                self.ebawe_project_list.append([self.ws_ebawe.cell(column=5, row=row_index).value, row_index])
                if proj_counter != 0:  # intentional omission of the loop operation in the first iteration
                    self.ebawe_project_list[proj_counter - 1].append(
                        self.ebawe_project_list[proj_counter][1] - self.ebawe_project_list[proj_counter - 1][1] - 6
                    )
                proj_counter += 1
        row_of_last_proj = self.ebawe_project_list[-1][1]
        number_of_elements_in_last_proj = 0
        while True:
            if self.ws_ebawe.cell(column=3, row=row_of_last_proj + number_of_elements_in_last_proj + 3).value is None:
                break
            number_of_elements_in_last_proj += 1
        self.ebawe_project_list[-1].append(number_of_elements_in_last_proj)

    def create_daily_report(self):
        """ Making a daily report - copy data from ebawe report to daily report,
        but area of elements are taken from yearly report. Also formatting daily report.
        :return:
        """
        # Heading
        self.ws_daily['H8'] = f"{self.user_data['dd']}.{self.user_data['mm']}.{self.user_data['yyyy']}"
        self.ws_daily['H10'] = f"E{self.user_data['ebawe']}"

        # Reducing row height
        small_rows = (4, 5, 6, 7, 9, 11)
        for row in small_rows:
            self.ws_daily.row_dimensions[row].height = 1

        # For loop for every project in self.ebawe_project_list
        for proj in self.ebawe_project_list:
            # Project headers ("Zlecenie")
            self.ws_daily.cell(row=proj[1], column=1).value = "Zlecenie:"
            self.ws_daily.cell(row=proj[1], column=5).value = proj[0]
            self.ws_daily.cell(row=proj[1], column=1).font = font_bold
            self.ws_daily.cell(row=proj[1], column=5).font = font_bold
            for row in self.ws_daily.iter_rows(min_row=proj[1], max_row=proj[1], min_col=1, max_col=9):
                for cell in row:
                    cell.fill = grey_fill
            # Row height reduction
            self.ws_daily.row_dimensions[proj[1] + 1].height = 4
            self.ws_daily.row_dimensions[proj[1] + proj[2] + 3].height = 4
            # Columns titles
            col_number = 0
            for title in self.title_tuple:
                self.ws_daily.cell(row=proj[1] + 2, column=self.title_index[col_number]).value = title
                col_number += 1
            for row in self.ws_daily.iter_rows(min_row=proj[1] + 2, max_row=proj[1] + 2, min_col=1, max_col=9):
                for cell in row:
                    cell.fill = title_fill
            # Sum of elements
            self.ws_daily.cell(row=proj[1] + proj[2] + 4, column=3).value = proj[2]
            self.ws_daily.cell(row=proj[1] + proj[2] + 4, column=3).border = thick_border
            self.ws_daily.cell(row=proj[1] + proj[2] + 4, column=3).fill = summary_fill
            # Finding proper column index in "yearly report"
            for cell in self.row_1_yearly:
                if cell.value == proj[0]:
                    col_index_in_yearly = cell.column
            try:
                col_index_in_yearly += 0
            except NameError:
                warn = f"Nie znaleziono odpowiedniego projektu\nw Excelu z powierzchniami brutto.\n\n{proj[0]}\n\n" \
                       f"Skrypt zamknie się bez zapisywania żadnych zmian.\n\n" \
                       f"Sprawdź, czy numery projektów są identyczne\nwe wszystkich wejściowych arkuszach excel\n" \
                       f"i uruchom skrypt ponownie."
                warning_msg_box(warn)
            # Project description
            description = self.ws_yearly.cell(row=2, column=col_index_in_yearly).value.replace("\n", " - ")
            self.ws_daily.cell(row=proj[1], column=10).value = description
            proj.append(description)
            # Main data table
            # While loop for every prefab element in the project
            el = 0  # el - element counter in each project
            while el < proj[2]:
                area_found = 0
                try:
                    num = int(self.ws_ebawe.cell(row=proj[1]+el+3, column=3).value)
                except ValueError:
                    warn = f"Nr elementu:  {self.ws_ebawe.cell(row=proj[1]+el+3, column=3).value}\nz projektu:  " \
                           f"{proj[0]}\n\nnie jest liczbą.\nSkrypt zamknie się bez zapisywania żadnych zmian.\n" \
                           f"Sprawdź numery elementów i uruchom skrypt ponownie."
                    warning_msg_box(warn)
                self.ws_daily.cell(row=proj[1]+el+3, column=3).value = self.ws_ebawe.cell(
                    row=proj[1]+el+3, column=3).value
                self.ws_daily.cell(row=proj[1]+el+3, column=3).border = border
                self.ws_daily.cell(row=proj[1]+el+3, column=4).value = self.ws_ebawe.cell(
                    row=proj[1]+el+3, column=4).value
                self.ws_daily.cell(row=proj[1]+el+3, column=4).border = border
                self.ws_daily.cell(row=proj[1]+el+3, column=5).border = border
                self.ws_daily.cell(row=proj[1]+el+3, column=7).value = self.ws_ebawe.cell(
                    row=proj[1]+el+3, column=9).value
                self.ws_daily.cell(row=proj[1]+el+3, column=7).border = border
                self.ws_daily.cell(row=proj[1]+el+3, column=8).border = border
                self.ws_daily.cell(row=proj[1]+el+3, column=9).value = self.ws_ebawe.cell(
                    row=proj[1]+el+3, column=11).value
                self.ws_daily.cell(row=proj[1]+el+3, column=9).border = border
                self.ws_daily.cell(row=proj[1]+el+3, column=6).border = border
                if self.ws_yearly.cell(row=num + 8, column=col_index_in_yearly).value is not None:
                    test_color = str(self.ws_yearly.cell(row=num+8, column=col_index_in_yearly).fill.start_color.index)
                    if test_color[-6:] != 'FFFF00':
                        self.ws_daily.cell(row=proj[1]+el+3, column=6).value = self.ws_yearly.cell(
                            row=num+8, column=col_index_in_yearly).value
                        self.ws_yearly.cell(row=num+8, column=col_index_in_yearly).fill = yellow_fill
                    else:
                        warn = f"Próbujesz wpisać do raportów płytę,\nktóra już wcześniej była zaraportowana:\n\n" \
                               f"Projekt:  {proj[0]}\nNumer elementu:  {num}\n\n" \
                               f"Skrypt zamknie się bez zapisywania żadnych zmian.\n" \
                               f"Zweryfikuj błąd i uruchom skrypt ponownie."
                        warning_msg_box(warn)
                else:
                    # Looking for the area in the next 10 columns on the right
                    for index_increase in range(1, 10):
                        if self.ws_yearly.cell(row=num+8, column=col_index_in_yearly+index_increase).value is not None:
                            test_color2 = str(self.ws_yearly.cell(
                                row=num+8, column=col_index_in_yearly+index_increase).fill.start_color.index)
                            if test_color2[-6:] != 'FFFF00':
                                self.ws_daily.cell(row=proj[1]+el+3, column=6).value = \
                                    self.ws_yearly.cell(row=num+8, column=col_index_in_yearly+index_increase).value
                                self.ws_yearly.cell(row=num+8, column=col_index_in_yearly+index_increase).fill = \
                                    yellow_fill
                                self.ws_daily.cell(row=proj[1]+el+3, column=11).value = \
                                    self.ws_yearly.cell(row=num+8, column=col_index_in_yearly+index_increase).value
                                self.ws_daily.cell(row=proj[1]+el+3, column=10).value = \
                                    self.ws_yearly.cell(row=5, column=col_index_in_yearly+index_increase).value
                                self.ws_daily.cell(row=proj[1]+el+3, column=10).fill = pink_fill
                                self.ws_daily.cell(row=proj[1]+el+3, column=11).fill = pink_fill
                                area_found = 1
                                break
                            else:
                                warn = f"Próbujesz wpisać do raportów płytę,\nktóra już wcześniej była zaraportowana:" \
                                       f"\n\nProjekt:  {proj[0]}\nNumer elementu:  {num}\n\n" \
                                       f"Skrypt zamknie się bez zapisywania żadnych zmian.\n" \
                                       f"Zweryfikuj błąd i uruchom skrypt ponownie."
                                warning_msg_box(warn)
                    if area_found != 1:
                        warn = f"Nie można pobrać powierzchni\nelementu nr:  {num}\nz projektu:  {proj[0]}\n\n" \
                               f"Skrypt zamknie się bez zapisywania żadnych zmian.\n" \
                               f"Sprawdź poprawność numeru tego elementu\ni uruchom skrypt ponownie."
                        warning_msg_box(warn)
                el += 1
            # Creating a sum formula
            sum_list = []  # init list of row indexes with cells which need to be sum
            for element in range(0, proj[2]):
                sum_list.append(proj[1]+3+element)
            formula = "=0"
            for row_index_to_sum in sum_list:
                formula += f"+F{row_index_to_sum}"
            self.ws_daily.cell(row=proj[1]+proj[2]+4, column=6).value = formula
            self.ws_daily.cell(row=proj[1]+proj[2]+4, column=6).fill = summary_fill
            self.ws_daily.cell(row=proj[1]+proj[2]+4, column=6).border = thick_border

            # Additional Sums
            add_sum_dict = {}
            # For loop for initializing keys
            for col in self.ws_daily.iter_cols(min_col=10, max_col=10, min_row=proj[1]+3, max_row=proj[1]+proj[2]+2):
                for cell in col:
                    if cell.value is not None:
                        add_sum_dict[cell.value] = 0
            # For loop for set values to keys
            for col in self.ws_daily.iter_cols(min_col=10, max_col=10, min_row=proj[1]+3, max_row=proj[1]+proj[2]+2):
                for cell in col:
                    if cell.value is not None:
                        add_sum_dict[cell.value] += self.ws_daily.cell(row=cell.row, column=cell.column+1).value
            proj.append(add_sum_dict)
            for key in add_sum_dict.keys():
                additional_sums_written = 0
                for col in self.ws_daily.iter_cols(
                        min_col=10, max_col=10, min_row=proj[1]+3, max_row=proj[1]+proj[2]+2):
                    for cell in col:
                        if cell.value == key and additional_sums_written == 0:
                            self.ws_daily.cell(row=cell.row, column=cell.column+2).value = add_sum_dict[key]
                            self.ws_daily.cell(row=cell.row, column=cell.column+2).fill = summary_fill
                            self.ws_daily.cell(row=cell.row, column=cell.column).fill = summary_fill
                            additional_sums_written = 1

    def painting_monthly_and_yearly_reports(self):
        """ Painting headers on green and elements cell on yellow in monthly and yearly reports.
        :return:
        """
        # Painting green "pow_do_raportu"
        for col in self.ws_yearly.iter_cols(
                min_row=9,
                max_row=self.yearly_max_row,
                min_col=2,
                max_col=self.yearly_max_col):
            proj_done = 0
            for cell in col:
                if cell.value is not None:
                    proj_done = 1
                    break
            for cell in col:
                test_color3 = str(cell.fill.start_color.index)
                if cell.value is not None and test_color3[-6:] != 'FFFF00':
                    proj_done = 0
                    break
            if proj_done == 1:
                self.ws_yearly.cell(row=4, column=col[4].column).fill = green_fill
                self.ws_yearly.cell(row=5, column=col[5].column).fill = green_fill

        for proj in self.ebawe_project_list:
            stop_painting = 0
            for proj_in_year in self.row_1_yearly:
                if proj[0] == proj_in_year.value:
                    list_of_proj_to_paint = [proj_in_year]
                    proj_in_year_col = proj_in_year.column
                    for index_increase in range(1, 11):
                        if self.ws_yearly.cell(row=1, column=proj_in_year_col + index_increase).value is None:
                            list_of_proj_to_paint.append(self.ws_yearly.cell(
                                row=1, column=proj_in_year_col+index_increase))
                        else:
                            break
                    len_to_paint = len(list_of_proj_to_paint)
                    paint_or_not = 1
                    for cell_number in range(0, len_to_paint):
                        test_color4 = str(self.ws_yearly.cell(
                            row=4, column=list_of_proj_to_paint[cell_number].column).fill.start_color.index)
                        if test_color4[-6:] != '548235':
                            paint_or_not = 0
                            break
                    if paint_or_not == 1:
                        for proj_in_month in self.row_1_monthly:
                            if proj[0] == proj_in_month.value:
                                proj_in_month_col = proj_in_month.column
                        for col_counter in range(0, len_to_paint):
                            self.ws_yearly.cell(row=1, column=proj_in_year_col+col_counter).fill = green_fill
                            self.ws_yearly.cell(row=2, column=proj_in_year_col+col_counter).fill = green_fill
                            self.ws_yearly.cell(row=3, column=proj_in_year_col+col_counter).fill = green_fill
                            # Painting yellow "month report"
                            try:
                                if col_counter == 0:
                                    self.ws_monthly.cell(row=1, column=proj_in_month_col).fill = yellow_fill
                                    self.ws_monthly.cell(row=2, column=proj_in_month_col).fill = yellow_fill
                                    self.ws_monthly.cell(row=3, column=proj_in_month_col).fill = yellow_fill
                                    self.ws_monthly.cell(row=4, column=proj_in_month_col).fill = yellow_fill
                                else:
                                    if self.ws_monthly.cell(
                                            row=1, column=proj_in_month_col + col_counter).value is not None or \
                                            self.ws_monthly.cell(
                                                row=2, column=proj_in_month_col + col_counter).value is None:
                                        stop_painting += 1
                                    if stop_painting == 0:
                                        self.ws_monthly.cell(
                                            row=2, column=proj_in_month_col+col_counter).fill = yellow_fill
                                        self.ws_monthly.cell(
                                            row=3, column=proj_in_month_col+col_counter).fill = yellow_fill
                            except NameError:
                                warn = f"Nie znaleziono odpowiedniego projektu\nw Excelu z miesięcznym raportem:\n\n" \
                                       f"{proj[0]}\n\nSkrypt zamknie się bez zapisywania żadnych zmian.\n\n" \
                                       f"Sprawdź, czy projekt znajduje się w tabeli\nz miesięcznym raportem " \
                                       f"i czy jest poprawnie wpisany.\nNastępnie uruchom skrypt ponownie."
                                warning_msg_box(warn)

        # Saving temporary file
        self.wb_daily.save(self.config.TEMPORARY_FILE)

    def filling_monthly_report(self):
        """ Filling month report and ExcelCompiler stuff which calculate value from formula
        :return:
        """
        excel_comp_obj = ExcelCompiler(filename=self.config.TEMPORARY_FILE)
        for proj in self.ebawe_project_list:
            col_index = "test"
            proj_name = proj[0]
            for cell in self.row_1_monthly:
                if cell.value == proj_name:
                    col_index = cell.column
            try:
                col_index += 0
            except (NameError, TypeError):
                os.remove(self.config.TEMPORARY_FILE)
                warn = f"Nie znaleziono odpowiedniego projektu\nw Excelu z miesięcznym raportem:\n\n{proj_name}\n\n" \
                       f"Skrypt zamknie się bez zapisywania żadnych zmian.\n\n" \
                       f"Sprawdź, czy projekt znajduje się w tabeli\n" \
                       f"z miesięcznym raportem i czy jest poprawnie wpisany.\nNastępnie uruchom skrypt ponownie."
                warning_msg_box(warn)
            row_index_to_evaluate = proj[1]+proj[2]+4
            evaluated_value = excel_comp_obj.evaluate(f"E{self.user_data['ebawe']}!F{row_index_to_evaluate}")
            # subtracting from the "evaluated value" values from the dictionary
            for dict_value in proj[4].values():
                evaluated_value -= dict_value
            # writing "evaluated value" in daily report
            for cell in self.row_1_yearly:
                if cell.value == proj_name:
                    col_index2 = cell.column
            if evaluated_value > 0.02:
                self.ws_daily.cell(row=row_index_to_evaluate, column=10).value = self.ws_yearly.cell(
                    row=5, column=col_index2).value
                self.ws_daily.cell(row=row_index_to_evaluate, column=11).value = evaluated_value
                self.ws_daily.cell(row=row_index_to_evaluate, column=10).fill = summary_fill
                self.ws_daily.cell(row=row_index_to_evaluate, column=11).fill = summary_fill
                self.ws_daily.cell(row=row_index_to_evaluate, column=10).border = border
                self.ws_daily.cell(row=row_index_to_evaluate, column=11).border = border
                # adding "evaluated value" to the dictionary
                proj[4][self.ws_yearly.cell(row=5, column=col_index2).value] = evaluated_value
            # actual completion of the monthly report
            for key in proj[4].keys():
                for product in range(0, 10):
                    if key == self.ws_monthly.cell(row=2, column=col_index+product).value and \
                            (self.ws_monthly.cell(row=1, column=col_index+product).value == proj[0] or
                             self.ws_monthly.cell(row=1, column=col_index + product).value is None):
                        self.ws_monthly.cell(row=int(self.user_data['dd'])+5,
                                             column=col_index+product).value = proj[4][key]
                        break

    def saving_all_reports_files(self):
        self.wb_month.save(self.config.MONTHLY_TEMP_PATH)
        self.wb_daily.save(self.config.DAILY_TEMP_PATH)
        self.wb_yearly.save(self.config.YEARLY_TEMP_PATH)
        os.remove(self.config.TEMPORARY_FILE)

        warning_msg_box("JUŻ  :)")

# endregion


if __name__ == "__main__":
    # Run welcome message box to take date and number of production line from user
    user_input = report_welcome_message_box()

    # Validate given data from user
    if user_input['dd'] == "" or user_input['mm'] == "" or user_input['yyyy'] == "" or user_input['ebawe'] == "":
        warning_msg = "Nie podano wszystkich danych w oknie startowym!\nKończę skrypt."
        warning_msg_box(warning_msg)
    elif not user_input['dd'].isnumeric() or not user_input['mm'].isnumeric() or not user_input['yyyy'].isnumeric():
        warning_msg = "Dane z datą w oknie startowym muszą być numeryczne!\nKończę skrypt."
        warning_msg_box(warning_msg)

    # Create config object
    # TODO: comment out the proper line from next two lines
    # config_obj = config_env['production'](dd=user_input['dd'],
    #                                       mm=user_input['mm'],
    #                                       yyyy=user_input['yyyy'],
    #                                       ebawe=user_input['ebawe'])
    config_obj = config_env['testing'](dd=user_input['dd'],
                                       mm=user_input['mm'],
                                       yyyy=user_input['yyyy'],
                                       ebawe=user_input['ebawe'])

    # Init and call Report object
    report_obj = Report(config=config_obj, user_data=user_input)
    report_obj()
