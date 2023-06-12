# Loading and filter data from PDF file to Excel

import fitz
import tkinter as tk
import sys
import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side


# Adjusting style
# color
no_color_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
# border
edge = Side(border_style='thin', color='000000')
border = Border(top=edge, bottom=edge, left=edge, right=edge)


# Hardcoded data and variables
safe_to_save = 1
column_found = 0
language = "empty"
# language = "PL"
# language = "ENG"
# file_path = "C:\Raporty\Rumia, Żeglarzy bud. 5 - strop nad parterem - Projekt.pdf"
# file_path = "C:\Raporty\Journalen HUS 1 CEILING PLAN 14.pdf"


# Welcome message box
ask_window = tk.Tk()
ask_window.title("inBet  Export area of elements")
ask_window.geometry("950x245")

ask_label = tk.Label(ask_window, text="Podaj numer projektu i jego wersję językową oraz bieżacy rok: ")
ask_label.pack(pady=5, padx=5)

ask_frame = tk.Frame(ask_window)
ask_frame.pack(pady=5, padx=5)

ask_proj_label = tk.Label(ask_frame, text="Projekt:")
ask_proj_label.grid(pady=2, padx=5, row=0, column=0)

ask_year_label = tk.Label(ask_frame, text="Rok:")
ask_year_label.grid(pady=2, padx=5, row=1, column=0)

ask_proj = tk.Entry(ask_frame, width=17)
ask_proj.grid(pady=2, padx=5, row=0, column=1)
ask_proj.insert(0, "111-1")

ask_year = tk.Entry(ask_frame, width=17)
ask_year.grid(pady=2, padx=5, row=1, column=1)
ask_year.insert(0, "2023")

ask_lang = tk.Listbox(ask_frame, width=7, height=2, selectmode=tk.SINGLE)
ask_lang.grid(pady=2, padx=5, row=0, column=2, rowspan=2)
ask_lang.insert(tk.END, "PL")
ask_lang.insert(tk.END, "ENG")

path_label = tk.Label(ask_window, text="\n\nPodaj ściężkę dostępu do pliku z projektem: ")
path_label.pack(pady=5, padx=5)

path_entry = tk.Entry(ask_window, width=150)
path_entry.pack(pady=5, padx=5)
path_entry.insert(0, "S:\Projekty\_RYSUNKI\FS\975 - WCK Sierakowice\975-2 - strop nad parterem\975-2 - Listy i zestawienia.pdf")


def button():
    global project_number, language, file_path, yyyy
    project_number = ask_proj.get()
    for i in ask_lang.curselection():
        language = ask_lang.get(i)
    file_path2 = path_entry.get()
    file_path = file_path2.replace("\"", "")
    yyyy = ask_year.get()
    if project_number != "" and language != "empty" and file_path != "" and yyyy != "":
        ask_window.destroy()


ask_button = tk.Button(ask_window, text="Zatwierdź", command=button, width=25, height=2)
ask_button.pack(pady=5, padx=5)

ask_window.mainloop()


# Warning message box
def warning(w):
    warning_window = tk.Tk()
    warning_window.title("Warning !")

    warning_label = tk.Label(warning_window, text=w)
    warning_label.pack(pady=20, padx=25)
    print(w)

    def warning_button():
        global safe_to_save
        safe_to_save = 0
        warning_window.destroy()
        sys.exit()
    warning_button = tk.Button(warning_window, text="OK", command=warning_button, width=25, height=2)
    warning_button.pack(pady=10, padx=15)

    warning_window.mainloop()


# Locations of the file
if os.path.isfile("S:\\DPP\\5_ZESTAWIENIA PREFABRYKACJI\\RAPORT PRODUKCJI FILIGRAN DZIENNIE\\Produkcja płyt wg projektów - " + yyyy + "_GR.xlsx"):
    pow_do_raportu = "S:\\DPP\\5_ZESTAWIENIA PREFABRYKACJI\\RAPORT PRODUKCJI FILIGRAN DZIENNIE\\Produkcja płyt wg projektów - " + yyyy + "_GR.xlsx"
else:
    pow_do_raportu = "S:\\DPP\\5_ZESTAWIENIA PREFABRYKACJI\\RAPORT PRODUKCJI FILIGRAN DZIENNIE\\Produkcja płyt wg projektów - " + yyyy + ".xlsx"


# Flags used to search for relevant data
if language == "PL":
    heading = "ZESTAWIENIE POJEDYNCZYCH PLYT"
    ceiling = "Strop filigr"
elif language == "ENG":
    heading = "S I N G L E   P A N E L   S C H E D U L E"
    ceiling = "Half floor"


# Loading PDF file
try:
    doc = fitz.open(file_path)
except (FileNotFoundError, NameError):
    war = """Nie znaleziono pliku z projektem pod wskazaną ścieżką dostepu.

Sprawdź poprawność lokalizacji pliku z projektem oraz czy plik istnieje."""
    warning(war)
full_text = ""
for page in doc:
    single_page = page.get_text()
    if heading in single_page:
        full_text += page.get_text()
# print(full_text)


# Creating and filling in the list of indexes under which the keyword appears (ceiling = keyword)
List_of_ceiling_index = []
current_ceiling_index = -5
full_text_2 = full_text         # full_text_2 is a working copy of full_text
len_of_full_text_2 = len(full_text_2)
while len_of_full_text_2 > 0 and current_ceiling_index != -1:
    current_ceiling_index = full_text_2.rfind(ceiling)
    List_of_ceiling_index.append(current_ceiling_index)
    full_text_2 = full_text_2[0:current_ceiling_index]
    len_of_full_text_2 = len(full_text_2)
if -1 in List_of_ceiling_index:
    List_of_ceiling_index.remove(-1)
List_of_ceiling_index.reverse()
# print(List_of_ceiling_index)
if len(List_of_ceiling_index) == 0:
    war = """Nie udało się pobrać żadnych powierzchni elementów.

Możliwe przyczyny to:
- Zaciągnięty został niewłaściwy plik projektu - niezawierający tabeli z powierzchniami.
- Został wybrany zły język projektu.
- Plik PDF nie jest typu "NATIVE".
  PDF'y w formie zdjęć (nie mające możliwości zazaczania tekstu kursorem) nie są obsługiwane.
- Zmienił się szablon projektu np.: nie zawiera już słów kluczowych, po których skrypt szuka i filtruje dane
  lub są one inaczej rozmieszczone itp."""
    warning(war)


# Creating and filling in the Dict of {element number : area}
Dict_of_area = {}

for ceiling_index in List_of_ceiling_index:
    # element number
    for i in range(1, 4):
        current_index = ceiling_index - i
        if full_text[current_index] == "|":
            end_slab_index = current_index
    for i in range(4, 15):
        current_index = ceiling_index - i
        if full_text[current_index] == "|":
            start_slab_index = current_index
    try:
        slab_number = int(full_text[start_slab_index+1:end_slab_index-2].strip())
    except (ValueError):
        war = """W pobranych z projektu danych przynajmniej część numerów elementów prefabrykowanych nie jest liczbami.

Aby skrypt działał poprawnie numery elementów muszą być liczbami."""
        warning(war)

    # area
    for i in range(40, 51):
        current_index = ceiling_index + i
        if full_text[current_index] == "|":
            start_area_index = current_index
    for i in range(51, 63):
        current_index = ceiling_index + i
        if full_text[current_index] == "|":
            end_area_index = current_index
    try:
        slab_area = float(full_text[start_area_index+1:end_area_index].strip())
    except (ValueError):
        war = """W pobranych z projektu danych przynajmniej część powierzchni nie jest liczbami
lub separatorem dziesiętnym jest przecinek zamiast kropki."""
        warning(war)
    Dict_of_area[slab_number] = slab_area

for i in Dict_of_area:
    print(i, Dict_of_area[i])


# Load workbooks and sheets
try:
    wb_pow_do_raportu = openpyxl.load_workbook(pow_do_raportu)
except (FileNotFoundError, NameError):
    war = """Nie znaleziono pliku excel z powierzchniami brutto.

Skrypt zamknie się bez zapisywania żadnych zmian.

Sprawdź czy plik jest stworzony.
Zweryfikuj jego nazwę, rozszerzenie, lokalizację itp.
i uruchom skrypt ponownie."""
    warning(war)
# Get workbook active sheet object from the active attribute or sheet name.
sheet_pow_do_raportu = wb_pow_do_raportu.active
# row 1 variables
row_1_pow_do_raportu = sheet_pow_do_raportu[1]


# Finding the proper column in Excel file
for cell in row_1_pow_do_raportu:
    if cell.value == project_number:
        column_index = cell.column
        column_found = 1
if column_found == 0:
    war = """Nie znaleziono wskazanego projektu w akruszu excel
z powierzchniami brutto.

Dodaj kolumnę z tym projektem do arkusza excel
i/lub
zweryfikuj czy poprownie został wpisany jego numer."""
    warning(war)


# Filling Excel column with data from Dict_of_area
for element in Dict_of_area:
    if sheet_pow_do_raportu.cell(row=element+8, column=column_index).value is None:
        sheet_pow_do_raportu.cell(row=element+8, column=column_index).value = Dict_of_area[element]
        sheet_pow_do_raportu.cell(row=element+8, column=column_index).border = border
        sheet_pow_do_raportu.cell(row=element+8, column=column_index).fill = no_color_fill
    else:
        war = """Niniejszy projekt jest już przynajmniej w części uzupełniony.

Aby nie nadpisać danych skrypt nie dokona ich wprowadzenia do excela.
Upewnij się również, czy wprowadzasz odpowiedni projekt."""
        warning(war)

# Saving file
if safe_to_save == 1:
    wb_pow_do_raportu.save("S:\\DPP\\5_ZESTAWIENIA PREFABRYKACJI\\RAPORT PRODUKCJI FILIGRAN DZIENNIE\\Produkcja płyt wg projektów - " + yyyy + "_GR.xlsx")
    warning("JUŻ  :)")
